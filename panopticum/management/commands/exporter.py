#!/usr/bin/env python

from __future__ import print_function

import logging
import os
import sys
import traceback
import warnings
from collections import OrderedDict

import openpyxl
import pytz
from django.core.management import BaseCommand
from openpyxl import *
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import WorksheetProperties
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from panopticum.models import *

warnings.simplefilter('ignore')


class ColRef(object):
    START_CELL = 'startcell'
    END_CELL = 'endcell'

    def __init__(self, tablename: str, column_settings):
        tablename.strip()
        self.sheet_name = tablename  # table.get_tablename()
        self.startcell = column_settings.get_sheet_metadata()[ColRef.START_CELL]
        self.endcell = column_settings.get_sheet_metadata()[ColRef.END_CELL]


class ColFormatter(object):
    DEFAULT_WIDTH = 10
    DEFAULT_WRAP = True

    def __init__(self, column_nm: str, **args):
        argkeys = args.keys()
        self.column_nm = column_nm
        self.width = args["width"] if "width" in argkeys else ColFormatter.DEFAULT_WIDTH
        self.wrap = args["wrap"] if "wrap" in argkeys else ColFormatter.DEFAULT_WRAP
        self.comment = args["comment"] if "comment" in argkeys else None
        self.reference: ColRef = args[
            "reference"] if "reference" in argkeys else None
        self.dataref: dict = args["sheet_metadata"] if "sheet_metadata" in argkeys else None
        if self.dataref is None:
            self.dataref = {ColRef.START_CELL: '$B$2', ColRef.END_CELL: '$B$2'}

    def update_sheet_metadata(self, metadata: dict):
        """ Helper function that holds references within sheet for this column
        e.g. data starting cell, data ending cell etc"""
        self.dataref = metadata

    def get_sheet_metadata(self):
        return self.dataref

    @classmethod
    def get_defaultcolumn(cls, column: str):
        return cls(column, width=ColFormatter.DEFAULT_WIDTH, wrap=ColFormatter.DEFAULT_WRAP)


class TableFormatter(object):
    """ Provides default formatting as well as provides registered class formatting
    """
    # Below COLUMN_* are assumed to be default columns each tables would have.
    # Non-required columns wouldn't be used hence no harm.
    COLUMN_ID = 'id'
    COLUMN_NM = 'name'
    COLUMN_DESC = 'description'
    COLUMN_ORDER = 'order'

    # TODO: HG:
    ''' a
        1. Check how we can use various Worksheet properties
            # https://openpyxl.readthedocs.io/en/stable/worksheet_properties.html
            >>> wsprops = ws.sheet_properties
            >>> wsprops.tabColor = "1072BA"
            >>> wsprops.filterMode = False
            >>> wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
            >>> wsprops.outlinePr.summaryBelow = False
            >>> wsprops.outlinePr.applyStyles = True
            >>> wsprops.pageSetUpPr.autoPageBreaks = True
            
        2. We also need to add more data validation forumals to the sheet as explained at 
        https://www.contextures.com/xlDataVal07.html - These validations will be part of ColumnFormatter or best 
        would be ColumnReference '''

    def __init__(self, tablename: str, sheet_pos=None, sheet_properties: WorksheetProperties = None,
                 table: dict = None):  # TODO: HG: Remove dict and replace with array of ColumnFormatter
        tablename = tablename.strip()
        self.table_nm = tablename
        self.sheet_pos = sheet_pos  # Sheet positions are just indicative and not guaranteed.
        self.sheet_properties = sheet_properties
        if not table:
            table = TableFormatter.get_defaulttable()
        else:
            # Ensure all keys are in lower case without leading or trailing spaces.
            table1 = {}
            for (k, v) in table.items():
                k = k.strip()
                table1[k.lower()] = v
            table = table1
            # TODO: HG: have concrete attributes (strong typing) instead of working with dictionary
        self.table = table

    def reg_col(self, column_settings: ColFormatter = None):
        """ Registers column settings like width value, wrap enabled?, if data is referenced in other sheets
        Method chaining pattern used """
        # TODO: HG: Use reference field
        if column_settings is None:
            raise ValueError("column_settings cannot be None")
        column_nm = column_settings.column_nm.strip()
        column_nm = column_nm.lower()
        column_settings.column_nm = column_nm
        self.table[column_nm] = column_settings
        return self

    def reg_cols_samesettings(self, col_names: list, column_setting: ColFormatter):
        """ Helper function to register multiple columns with same settings
        Follows Method chaining pattern
        """
        for column_nm in col_names:
            self.reg_col(ColFormatter(column_nm, width=column_setting.width, wrap=column_setting.wrap))
        return self

    def reg_cols(self, columns: list):
        """ Helper function to register multiple columns with same settings
        Follows Method chaining pattern
        """
        for column in columns:
            self.reg_col(column)
        return self

    def get_tablename(self):
        return self.table_nm

    def update_table(self, columns: [ColFormatter]):
        for column_setting in columns:
            column_nm = column_setting.get_columnname()
            column_nm = column_nm.strip()
            column_setting.column_nm = column_nm.lower()
            self.table[column_nm.lower()] = column_setting

    def get_column(self, col: str) -> ColFormatter:
        """ Returns column formatting with registered values else returns default """
        try:
            col = col.strip()
            return self.table[col.lower()]
        except KeyError:
            return ColFormatter.get_defaultcolumn(col.lower())

    @staticmethod
    def get_defaulttable() -> dict:
        """ Returns commonly available cell formatting """
        return {
            TableFormatter.COLUMN_ID: ColFormatter(TableFormatter.COLUMN_ID),
            TableFormatter.COLUMN_NM: ColFormatter(TableFormatter.COLUMN_NM, width=20),
            TableFormatter.COLUMN_DESC: ColFormatter(TableFormatter.COLUMN_DESC, width=50),
            TableFormatter.COLUMN_ORDER: ColFormatter(TableFormatter.COLUMN_ORDER)
        }

    def set_tablename(self, tablename):
        self.table_nm = tablename
        pass


class Utility:
    DEFAULT_AUTHOR = 'Admin'

    CMDOPTION_XLS_FILE = 'xls_file'
    CMDOPTION_OVERWRITE = 'overwrite'

    SHEETNM_CATEGORIES = 'CompCategories'
    SHEETNM_SUB_CATEGORIES = 'CompSubCategories'
    SHEETNM_PRIVACIES = 'DataPrivacies'
    SHEETNM_RUNTIME_TYPES = 'RuntimeTypes'
    SHEETNM_COUNTRIES = 'Countries'
    SHEETNM_DB_VENDORS = 'DBVendors'
    SHEETNM_DATACENTERS = 'DataCenters'
    SHEETNM_DEPL_ENVS = 'DeploymentEnvironment'
    SHEETNM_DEPL_LOCATIONS = 'DeploymentLocations'
    SHEETNM_LOGGERS = 'Loggers'
    SHEETNM_FRAMEWORKS = 'Frameworks'
    SHEETNM_LANGUAGES = 'Languages'
    SHEETNM_PROD_FAMILIES = 'ProductFamilies'
    SHEETNM_PROD_VERSIONS = 'ProductVersions'
    SHEETNM_APP_VENDORS = 'AppVendors'
    SHEETNM_PORTS = 'Ports'
    SHEETNM_DEPL_MODEL = 'Deployment Models'
    SHEETNM_COMPONENTS = 'Components'
    SHEETNM_COMPONENT_VERSIONS = 'CompVersions'
    SHEETNM_COMPONENT_DEPENDENCIES = 'CompDependencies'
    SHEETNM_ORM = 'ORMs'
    SHEETNM_COMPONENT_DEPLS = 'CompDeployments'
    SHEETNM_COMPONENT_VERSIONS_LIST = 'CompVersionList'
    SHEETNM_DEFAULTS = '_Defaults'

    sheet_settings_map = {  # IMP: This map only covers simple Models without references.
        # Check out below code for complex Models
        # Remember we have default values for id, name, description, order
        ComponentCategoryModel: TableFormatter(SHEETNM_CATEGORIES, sheet_pos=1),
        ComponentSubcategoryModel: TableFormatter(SHEETNM_SUB_CATEGORIES, sheet_pos=2).
            reg_col(ColFormatter("category", width=20)),
        ComponentDataPrivacyClassModel: TableFormatter(SHEETNM_PRIVACIES, sheet_pos=3),
        ComponentRuntimeTypeModel: TableFormatter(SHEETNM_RUNTIME_TYPES, sheet_pos=4),
        CountryModel: TableFormatter(SHEETNM_COUNTRIES, sheet_pos=5),
        DatabaseVendorModel: TableFormatter(SHEETNM_DB_VENDORS, sheet_pos=6),
        DatacenterModel: TableFormatter(SHEETNM_DATACENTERS, sheet_pos=7)
            .reg_cols([ColFormatter("info", width=50),
                       ColFormatter("grafana", width=50),
                       ColFormatter("metrics", width=50)]),
        DeploymentEnvironmentModel: TableFormatter(SHEETNM_DEPL_ENVS, sheet_pos=8),
        LoggerModel: TableFormatter(SHEETNM_LOGGERS, sheet_pos=9),
        FrameworkModel: TableFormatter(SHEETNM_FRAMEWORKS, sheet_pos=10)
            .reg_col(ColFormatter("language")),
        ProgrammingLanguageModel: TableFormatter(SHEETNM_LANGUAGES, sheet_pos=11),
        ProductFamilyModel: TableFormatter(SHEETNM_PROD_FAMILIES, sheet_pos=12),
        SoftwareVendorModel: TableFormatter(SHEETNM_APP_VENDORS, sheet_pos=13),
        TCPPortModel: TableFormatter(SHEETNM_PORTS, sheet_pos=14),
        DeploymentLocationClassModel: TableFormatter(SHEETNM_DEPL_LOCATIONS, sheet_pos=15)
            .reg_col(ColFormatter("shortname", width=15)),
        ComponentVersionListModel: TableFormatter(SHEETNM_COMPONENT_VERSIONS_LIST, sheet_pos=16)
            .reg_cols([ColFormatter("version", width=15),
                       ColFormatter("release_time", width=20)])
    }

    parser = None

    @staticmethod
    def register_parser(p):
        Utility.parser = p

    @staticmethod
    def usage():
        Utility.parser.print_help()


class XlsDocument:
    workbook: Workbook

    def __init__(self, filename, overwrite=False,
                 enable_registry=False):  # TODO: HG: Remove Registry code (though works) if not used in near future.
        self._filename = filename
        self.workbook = openpyxl.Workbook()
        self._registry = {}  # Keeps track of all the worksheets XlsDocument has worked upon so far
        self._registry_flag = enable_registry

        if os.path.isfile(self._filename) and overwrite is False:
            raise FileExistsError("%s [%s] file already exists without overwrite flag", self._filename)
        self.workbook = openpyxl.Workbook()
        self.workbook.save(self._filename)  # Can raise PermissionError

    def get_sheet_from_registry(self, sheet_name):
        if self._registry_flag is False:
            raise PermissionError("Registry is not enabled.")

        if sheet_name in self._registry.keys():
            return self._registry[sheet_name]
        else:
            raise ValueError("Registry doesn't have value for [%s]", sheet_name)

    def _update_sheet(self, sheet_details: TableFormatter, data):
        sheet_name = sheet_details.get_tablename()
        logging.debug('Creating/Updating sheet_name [%s]', sheet_name)
        if len(data) <= 0:
            logging.error('No values to insert for [%s]', sheet_name)
            return

        column_nms = self._get_column_names(data[0])
        if column_nms is None:
            logging.error("Error occured in retrieving column names. Exiting!")
        sheet = self._get_sheet_by_name(sheet_details, column_nms)
        row_no = 1
        for row in data:
            row_no += 1
            if isinstance(row, dict):
                sheet.append([v for k, v in row.items()])
            else:
                sheet.append([v for k, v in vars(row).items() if k != '_state'])

        # Adjusting column and cell settings.
        count = 1
        column_letter = 'A'
        for col in column_nms:
            setting = sheet_details.get_column(col)
            if count <= 26:
                column_letter: str = chr(64 + count)
            else:
                column_letter: str = chr(64 + int(count / 26)) + chr(
                    64 + (int(count % 26) if int(count % 26) != 0 else 1))
            XlsDocument._style_column(sheet, column_letter, setting)
            count += 1

        last_column = column_letter

        # Add table style to the data
        medium_style: TableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                                               showRowStripes=True,
                                                                               showLastColumn=False)
        table = openpyxl.worksheet.table.Table(ref="A1:{0}{1}".format(last_column, row_no),
                                               displayName=sheet_name.replace(" ", ""),
                                               tableStyleInfo=medium_style)

        # Other Worksheet level settings
        sheet.alignment = Alignment(horizontal="justify", wrapText=True)
        sheet.freeze_panes = "A2"
        # try out what is the value shown by sheet.dimensions

        sheet.add_table(table)
        if self._registry_flag is True:
            self._registry[sheet_details.get_tablename()] = sheet
        return

    def bulk_update(self, dict_data: OrderedDict):
        logging.debug('Input parameters [%s]', dict_data)
        for item, values in dict_data.items():
            self._update_sheet(item, values)

        self.workbook.save(self._filename)

    def update_sheet(self, sheet_details: TableFormatter, data):
        """ save sheet object if registry object is provided """
        self._update_sheet(sheet_details, data)
        self.workbook.save(self._filename)

    def get_cell_coordinates(self, sheet_name, column_name, start_nm='startcell', end_nm='endcell'):
        """ This function will strip all trailing spaces and perform case insensitive search on column_name """
        column_name = column_name.strip()
        sheet = self._registry[sheet_name] if sheet_name in self._registry.keys() else None
        if sheet is None:
            raise KeyError("Sheet name [{0}] doesn't exist in registry".format(sheet_name.lower()))

        for cell in sheet[1]:
            if cell.value.lower() == column_name.lower():
                return {
                    start_nm: "${0}$2".format(cell.column_letter),
                    end_nm: "${0}${1}".format(cell.column_letter, sheet.max_row if sheet.max_row > 1 else 2)
                }

        raise KeyError("Column name [{0}] doesn't exist for Sheet name [{1}]".format(column_name, sheet_name))

    @staticmethod
    def _get_column_names(data):
        if isinstance(data, dict):
            return data.keys()
        else:
            try:
                col_names = vars(data).copy()
                col_names.pop('_state')
                return col_names.keys()
            except KeyError as e:
                logging.error("error retrieving [_state] key. Exception [%s]", e)
                return None

    @staticmethod
    def _style_column(sheet: Worksheet, column_letter: str, column_details: ColFormatter):
        """ Helper function that will wrap all the cells in column
        NOTE: No data validation done hence any exceptions should be handled by caller. """
        sheet.column_dimensions[column_letter].width = column_details.width
        # if column_details.comment is not None:
        #     sheet["${0}$1".format(column_letter)].comment = column_details.comment
        cr = column_details.reference
        if cr is not None:
            dv = DataValidation(type="list",
                                formula1="{0}!{1}:{2}".format(quote_sheetname(cr.sheet_name),
                                                              cr.startcell,
                                                              cr.endcell))
            dv.add('{0}2:{0}{1}'.format(column_letter, len(sheet[column_letter])))
            sheet.add_data_validation(dv)
        col_metadata = dict(startcell="${0}$2".format(column_letter),
                            endcell="${0}${1}".format(column_letter, len(sheet[column_letter])))
        column_details.update_sheet_metadata(col_metadata)
        if column_details.wrap is True:
            for cell in sheet[column_letter]:
                cell.alignment = Alignment(wrapText=True)
                if column_details.comment is not None:
                    cell.comment = column_details.comment

    def _get_sheet_by_name(self, sheet_details: TableFormatter, column_nms):
        sheet_name = sheet_details.get_tablename()
        try:
            sheet = self.workbook.get_sheet_by_name('Sheet')
            sheet.title = sheet_name
        except KeyError:
            try:
                sheet = self.workbook[sheet_name]
            except KeyError:
                sheet = self.workbook.create_sheet(title=sheet_name, index=sheet_details.sheet_pos)
                if sheet_details.sheet_properties is not None:
                    sheet.sheet_properties = sheet_details.sheet_properties

        sheet.append(list(column_nms))
        return sheet


class Exporter(object):  # TODO: HG: Create this as abstract base class using ABC
    xlsdoc: XlsDocument

    def __init__(self, xlsdoc):
        self.xlsdoc = xlsdoc
        pass

    # @abstractmethod
    def export(self):
        pass

    @staticmethod
    def _get_multi_item_string(iterables):
        tmp_str: str = ""
        if isinstance(iterables, list):
            for i in iterables:
                tmp_str += "* " + i + "\n"
        else:
            for d in iterables.all() if iterables is not None else []:
                tmp_str += "* " + d.name + "\n"

        return tmp_str[:-1]  # remove last character '\n'

    @staticmethod
    def _get_printable_person_info(person: User):
        return (person.first_name + " " + person.last_name + " <" + person.email + ">") if person is not None else ""

    def _get_printable_people_info(self, people):
        tmp_str: str = ""
        for d in people.all() if people is not None else []:
            tmp_str += "* " + self._get_printable_person_info(d) + ";" + "\n"

        return tmp_str[:-2]  # remove last character ','


# Used with the all the simple table exports
class GenericExporter(Exporter):
    classname = None

    def __init__(self, xlsdoc, classname):
        super().__init__(xlsdoc)
        self.classname = classname

    def export(self):
        logging.debug('Exporting for [%s]', self.classname)

        if self.classname not in Utility.sheet_settings_map.keys():
            raise KeyError("Sheet name not found for class_name [{0}]".format(self.classname))

        self.xlsdoc.update_sheet(Utility.sheet_settings_map[self.classname],
                                 self.classname.objects.all())
        pass


class ComponentSubcategoryExporter(Exporter):
    def __init__(self, xlsdoc):
        super().__init__(xlsdoc)

    @classmethod
    def get_sheet_details(cls, override_original=True) -> TableFormatter:
        sheet_details = TableFormatter(Utility.SHEETNM_SUB_CATEGORIES, 1) \
            .reg_col(ColFormatter("category",
                                  width=20,
                                  reference=ColRef(
                                      Utility.SHEETNM_CATEGORIES,
                                      Utility.sheet_settings_map[ComponentCategoryModel].get_column("name"))
                                  )
                     )
        if override_original is True:
            Utility.sheet_settings_map[ComponentSubcategoryModel] = sheet_details
        return sheet_details

    def export(self):
        logging.debug('Exporting for [%s]', ComponentSubcategoryModel)
        subcategories = []
        tmp_subcategories = ComponentSubcategoryModel.objects.all()
        logging.debug('Received [%d] records', len(tmp_subcategories))
        for subcategory_obj in tmp_subcategories:
            tmp = dict(id=subcategory_obj.id, name=subcategory_obj.name, category=subcategory_obj.category.name,
                       description=subcategory_obj.description)
            subcategories.append(tmp)

        self.xlsdoc.update_sheet(ComponentSubcategoryExporter.get_sheet_details(), subcategories)


class FrameworkExporter(Exporter):
    def __init__(self, xlsdoc):
        super().__init__(xlsdoc)

    @classmethod
    def get_sheet_details(cls, override_original=True) -> TableFormatter:
        sheet_details = TableFormatter(Utility.SHEETNM_FRAMEWORKS, 1) \
            .reg_col(ColFormatter("language",
                                  width=20,
                                  reference=ColRef(Utility.SHEETNM_LANGUAGES,
                                                   Utility.sheet_settings_map[ProgrammingLanguageModel].get_column(
                                                       "name")))
                     )
        if override_original is True:
            Utility.sheet_settings_map[FrameworkModel] = sheet_details
        return sheet_details

    def export(self):
        logging.debug('Exporting for [%s]', FrameworkModel)
        frameworks = []
        tmp_framework_objs = FrameworkModel.objects.all()
        logging.debug('Received [%d] records', len(tmp_framework_objs))
        for framework_obj in tmp_framework_objs:
            tmp = dict(id=framework_obj.id, name=framework_obj.name, language=framework_obj.language.name)
            frameworks.append(tmp)

        self.xlsdoc.update_sheet(FrameworkExporter.get_sheet_details(), frameworks)


class ORMExporter(Exporter):
    def __init__(self, xlsdoc):
        super().__init__(xlsdoc)

    @classmethod
    def get_sheet_details(cls, override_original=True) -> TableFormatter:
        sheet_details = TableFormatter(Utility.SHEETNM_ORM, 1) \
            .reg_col(ColFormatter("language",
                                  width=20,
                                  reference=ColRef(Utility.SHEETNM_LANGUAGES,
                                                   Utility.sheet_settings_map[ProgrammingLanguageModel].get_column(
                                                       "name"))

                                  )
                     )
        if override_original is True:
            Utility.sheet_settings_map[ORMModel] = sheet_details
        return sheet_details

    def export(self):
        logging.debug('Exporting for [%s]', ORMModel)
        orms = []
        tmp_orm_objs = ORMModel.objects.all()
        logging.debug('Received [%d] records', len(tmp_orm_objs))
        for orm_obj in tmp_orm_objs:
            tmp = dict(id=orm_obj.id, name=orm_obj.name, language=orm_obj.language.name)
            orms.append(tmp)

        self.xlsdoc.update_sheet(ORMExporter.get_sheet_details(), orms)


class ProductVersionExporter(Exporter):
    def __init__(self, xlsdoc):
        super().__init__(xlsdoc)

    @classmethod
    def get_sheet_details(cls, override_original=True) -> TableFormatter:
        sheet_details = TableFormatter(Utility.SHEETNM_PROD_VERSIONS, 1) \
            .reg_col(ColFormatter("family",
                                  width=20,
                                  reference=ColRef(Utility.SHEETNM_PROD_FAMILIES,
                                                   Utility.sheet_settings_map[ProductFamilyModel].get_column(
                                                       "name"))
                                  )
                     )
        if override_original is True:
            Utility.sheet_settings_map[ProductVersionModel] = sheet_details
        return sheet_details

    def export(self):
        logging.debug('Exporting for [%s]', ProductVersionModel)
        prod_versions = []
        tmp_prod_versions = ProductVersionModel.objects.all()
        logging.debug('Received [%d] records', len(tmp_prod_versions))
        for provider_obj in tmp_prod_versions:
            tmp = dict(id=provider_obj.id, shortname=provider_obj.shortname, name=provider_obj.name,
                       family=provider_obj.family.name, order=provider_obj.order)
            prod_versions.append(tmp)

        self.xlsdoc.update_sheet(ProductVersionExporter.get_sheet_details(), prod_versions)


class ComponentDeploymentExporter(Exporter):
    def __init__(self, xlsdoc):
        super().__init__(xlsdoc)

    @classmethod
    def get_sheet_details(cls, override_original=True) -> TableFormatter:
        sheet_details = TableFormatter(Utility.SHEETNM_COMPONENT_DEPLS, sheet_pos=3,
                                       sheet_properties=WorksheetProperties(tabColor="00008B")) \
            .reg_cols_samesettings(["service_name", "binary_name", "open_ports"], ColFormatter("", width=20)) \
            .reg_cols([
            ColFormatter("location_class", width=20,
                         reference=ColRef(Utility.SHEETNM_DEPL_LOCATIONS,
                                          Utility.sheet_settings_map[DeploymentLocationClassModel].get_column(
                                              "name"))),
            ColFormatter("product_version", width=20,
                         reference=ColRef(Utility.SHEETNM_PROD_VERSIONS,
                                          Utility.sheet_settings_map[ProductVersionModel].get_column("name"))),
            ColFormatter("environment", width=20,
                         reference=ColRef(Utility.SHEETNM_DEPL_ENVS,
                                          Utility.sheet_settings_map[DeploymentEnvironmentModel].get_column(
                                              "name"))),
            ColFormatter("component_version", width=25,
                         comment=Comment(text="Make sure such entry matches in either of 'CompVersion' sheets."
                                              "Allowed entry is '<component name> - <component version>'."
                                              "e.g. File Service - 1.0",
                                         author=Utility.DEFAULT_AUTHOR,
                                         height=110, width=230)
                         # reference=ColRef(Utility.SHEETNM_COMPONENT_VERSIONS,
                         #                  Utility.sheet_settings_map[ComponentVersionModel].get_column(
                         #                      "version"))
                         ),
            ColFormatter("open_ports", width=20,
                         comment=Comment(
                             text="Allows multiple entries in a cell. "
                                  "Prefix each entry with '* ' and end with end of line. "
                                  "Each entry should match `name` in sheet {0}. "
                                  "e.g. "
                                  "* HTTP-8081".format(Utility.SHEETNM_PORTS),
                             author=Utility.DEFAULT_AUTHOR,
                             height=110, width=230)
                         # reference=ColRef(Utility.SHEETNM_PORTS,
                         #                  Utility.sheet_settings_map[TCPPortModel].get_column(
                         #                      "name"))
                         ),
        ])
        if override_original is True:
            Utility.sheet_settings_map[ComponentDeploymentModel] = sheet_details
        return sheet_details

    # def get_portdetails(self, ports):
    #     portlst= []
    #     for p in ports if ports is not None else []:
    #         portlst.append('{0}: {1}'.format(p.name, p.port))
    #     return portlst

    def export(self):
        logging.debug('Exporting for [%s]', ComponentDeploymentModel)
        comp_depls = []
        tmp_comp_depls = ComponentDeploymentModel.objects.all()
        logging.debug('Received [%d] records', len(tmp_comp_depls))
        for comp_depl_obj in tmp_comp_depls:
            tmp = dict(id=comp_depl_obj.id, name=comp_depl_obj.name,
                       component_version="{0}".format(str(comp_depl_obj.component_version)),
                       location_class=comp_depl_obj.location_class.name,
                       product_version=comp_depl_obj.product_version.name, environment=comp_depl_obj.environment.name,
                       service_name=comp_depl_obj.service_name, binary_name=comp_depl_obj.binary_name,
                       open_ports=super()._get_multi_item_string(comp_depl_obj.open_ports),
                       notes=comp_depl_obj.notes
                       )
            comp_depls.append(tmp)
        del tmp_comp_depls

        self.xlsdoc.update_sheet(ComponentDeploymentExporter.get_sheet_details(), comp_depls)


class ComponentExporter(Exporter):
    def __init__(self, xlsdoc):
        super().__init__(xlsdoc)

    @classmethod
    def get_sheet_details(cls, override_original=True) -> TableFormatter:
        sheet_details = TableFormatter(Utility.SHEETNM_COMPONENTS, sheet_pos=0,
                                       sheet_properties=WorksheetProperties(tabColor="00008B")) \
            .reg_cols([
            ColFormatter("life_status", width=10),
            ColFormatter("runtime_type", width=13,
                         reference=ColRef(Utility.SHEETNM_RUNTIME_TYPES,
                                          Utility.sheet_settings_map[
                                              ComponentRuntimeTypeModel].get_column(
                                              "name"))),
            ColFormatter("data_privacy_class", width=13,
                         reference=ColRef(Utility.SHEETNM_PRIVACIES,
                                          Utility.sheet_settings_map[
                                              ComponentDataPrivacyClassModel].get_column(
                                              "name"))),
            ColFormatter("category", width=15,
                         reference=ColRef(Utility.SHEETNM_CATEGORIES,
                                          Utility.sheet_settings_map[ComponentCategoryModel].get_column("name"))),
            ColFormatter("subcategory", width=15, reference=ColRef(Utility.SHEETNM_SUB_CATEGORIES,
                                                                   Utility.sheet_settings_map[
                                                                       ComponentSubcategoryModel].get_column(
                                                                       "name"))),
            ColFormatter("vendor", width=13, reference=ColRef(Utility.SHEETNM_APP_VENDORS,
                                                              Utility.sheet_settings_map[
                                                                  SoftwareVendorModel].get_column("name")))
        ])

        if override_original is True:
            Utility.sheet_settings_map[ComponentModel] = sheet_details

        return sheet_details

    def export(self):
        logging.debug('Exporting for [%s]', ComponentModel)
        components = []
        tmp_components = ComponentModel.objects.all()
        logging.debug('Received [%d] records', len(tmp_components))
        for comp_obj in tmp_components:
            tmp = {'id': comp_obj.id, 'name': comp_obj.name, 'description': comp_obj.description,
                   'life_status': comp_obj.life_status, 'runtime_type': comp_obj.runtime_type.name,
                   'data_privacy_class': comp_obj.data_privacy_class.name, 'category': comp_obj.category.name,
                   'subcategory': comp_obj.subcategory.name, 'vendor': comp_obj.vendor.name}
            components.append(tmp)
        del tmp_components

        self.xlsdoc.update_sheet(ComponentExporter.get_sheet_details(), components)


class ComponentVersionExporter(Exporter):
    def __latest_export(self, components):
        """ Component version objects with latest version only """
        pass

    def __all_wo_latest_export(self, components):
        """ Component version objects without latest version """
        pass

    @classmethod
    def get_sheet_details(cls, override_original=True):
        sheet_details = ComponentExporter.get_sheet_details(False)
        sheet_details.reg_cols([
            ColFormatter("component", width=25,
                         reference=ColRef(Utility.SHEETNM_COMPONENTS,
                                          Utility.sheet_settings_map[ComponentModel].get_column("name"))
                         ),
            ColFormatter("comments", width=30),
            ColFormatter("depends_on", width=25,
                         comment=Comment(
                             text="Allows multiple entries in a cell. Prefix each entry with '* ' and end with end of "
                                  "line. Each entry should match `name` in sheet {0}. e.g. "
                                  "* Resource Manager".format(Utility.SHEETNM_COMPONENTS),
                             author=Utility.DEFAULT_AUTHOR,
                             height=110, width=230),
                         #  Macros are not supported by openpyxl as of 13th Feb 2020 :( ... Below JFYI:
                         #  FYI - Below is macro for allowing multiple entries in a cell with data validation
                         #  validation Refer - https://trumpexcel.com/select-multiple-items-drop-down-list-excel/
                         #  reference=ColumnReference(Utility.SHEETNM_COMPONENTS,
                         #  startcell=self.xlsdoc.get_startcell( Utility.SHEETNM_COMPONENTS, "name"),
                         #  endcell=self.xlsdoc.get_endcell( Utility.SHEETNM_COMPONENTS, "name"))
                         ),
            ColFormatter("dev_language", width=20,
                         comment=Comment(
                             text="Allows multiple entries in a cell. Prefix each entry with '* ' and end with end of "
                                  "line. Each entry should match `name` in sheet {0}. e.g. "
                                  "* Python3.6".format(Utility.SHEETNM_LANGUAGES),
                             author=Utility.DEFAULT_AUTHOR,
                             height=110, width=230)
                         # reference=ColRef(Utility.SHEETNM_LANGUAGES,
                         #                  Utility.sheet_settings_map[ProgrammingLanguageModel].get_column("name"))
                         ),
            ColFormatter("dev_framework", width=20,
                         comment=Comment(
                             text="Allows multiple entries in a cell. Prefix each entry with '* ' and end with end of "
                                  "line. Each entry should match `name` in sheet {0}. e.g. "
                                  "* stored procs".format(Utility.SHEETNM_FRAMEWORKS),
                             author=Utility.DEFAULT_AUTHOR,
                             height=110, width=230)
                         # reference=ColRef(Utility.SHEETNM_FRAMEWORKS,
                         #                  Utility.sheet_settings_map[FrameworkModel].get_column("name"))
                         ),
            ColFormatter("dev_database", width=20,
                         comment=Comment(
                             text="Allows multiple entries in a cell. Prefix each entry with '* ' and end with end of "
                                  "line. Each entry should match `name` in sheet {0}. e.g. "
                                  "* Percona".format(Utility.SHEETNM_DB_VENDORS),
                             author=Utility.DEFAULT_AUTHOR,
                             height=110, width=230)
                         # reference=ColRef(Utility.SHEETNM_DB_VENDORS,
                         #                  Utility.sheet_settings_map[DatabaseVendorModel].get_column("name"))
                         ),
            ColFormatter("dev_orm", width=20,
                         comment=Comment(
                             text="Allows multiple entries in a cell. Prefix each entry with '* ' and end with end of "
                                  "line. Each entry should match `name` in sheet {0}. e.g. "
                                  "* Hibernate".format(Utility.SHEETNM_ORM),
                             author=Utility.DEFAULT_AUTHOR,
                             height=110, width=230)
                         # reference=ColRef(Utility.SHEETNM_ORM,
                         #                  Utility.sheet_settings_map[ORMModel].get_column("name"))
                         ),
            ColFormatter("dev_logging", width=20,
                         comment=Comment(
                             text="Allows multiple entries in a cell. Prefix each entry with '* ' and end with end of "
                                  "line. Each entry should match `name` in sheet {0}. e.g. "
                                  "* Logf".format(Utility.SHEETNM_LOGGERS),
                             author=Utility.DEFAULT_AUTHOR,
                             height=110, width=230)
                         # reference=ColRef(Utility.SHEETNM_LOGGERS,
                         #                  Utility.sheet_settings_map[LoggerModel].get_column("name"))
                         ),
            ColFormatter("meta_product_versions", width=20,
                         comment=Comment(
                             text="Allows multiple entries in a cell. Prefix each entry with '* ' and end with end of "
                                  "line. Each entry should match `name` in sheet {0}. e.g. "
                                  "* My Cloud Product 1.0".format(Utility.SHEETNM_PROD_VERSIONS),
                             author=Utility.DEFAULT_AUTHOR,
                             height=110, width=230)
                         # reference=ColRef(Utility.SHEETNM_PROD_VERSIONS,
                         #                  Utility.sheet_settings_map[ProductVersionModel].get_column("name"))
                         ),

        ]) \
            .reg_cols_samesettings(
            ["compliance_api_signoff", "compliance_fips_signoff", "compliance_gdpr_signoff", "meta_update_by",
             "mt_db_anonymisation_signoff", "mt_http_tracing_signoff", "mt_logging_completeness_signoff",
             "mt_logging_format_signoff", "mt_logging_sanitization_signoff", "mt_logging_storage_signoff",
             "op_alerts_signoff", "op_backup_signoff", "op_failover_signoff", "op_guide_signoff",
             "op_horizontal_scalability_signoff", "op_metrics_signoff", "op_scaling_guide_signoff",
             "op_sla_guide_signoff", "op_zero_downtime_signoff", "owner_maintainer", "owner_responsible_qa",
             "qa_anonymisation_tests_signoff", "qa_api_tests_signoff", "qa_e2e_tests_signoff",
             "qa_longhaul_tests_signoff", "qa_manual_tests_signoff", "qa_perf_tests_signoff",
             "qa_security_tests_signoff", "qa_unit_tests_signoff", "qa_upgrade_tests_signoff_id"],
            ColFormatter("", width=25, comment=Comment(text="Allows multiple entries in a cell."
                                                            "Each item should follow format 'firstname lastname "
                                                            "<emailid>'",
                                                       author=Utility.DEFAULT_AUTHOR,
                                                       height=110, width=230))
        ) \
            .reg_cols_samesettings(
            ["owner_product_manager", "owner_program_manager", "owner_escalation_list", "owner_expert",
             "owner_architect"],
            ColFormatter("", width=25, comment=Comment(
                text="Allows multiple entries in a cell. Prefix each entry with '* ' and end with end of line. "
                     "Each item should follow format 'firstname lastname <emailid>'."
                     "e.g. "
                     "* Alex Ken <Alex.Ken@company.com>",
                author=Utility.DEFAULT_AUTHOR,
                height=110, width=230))
        ) \
            .reg_cols_samesettings(
            ["dev_raml", "dev_repo", "dev_public_repo", "dev_docs", "dev_public_docs",
             "dev_commit_link"],
            ColFormatter("", width=25,
                         comment=Comment(text="Allows multiple entries in a cell. Separate each link with space",
                                         author=Utility.DEFAULT_AUTHOR,
                                         height=110, width=230))
        ) \
            .reg_cols_samesettings(["dev_jira_component", "dev_build_jenkins_job", "meta_profile_not_filled_fields",
                                    "meta_bad_rating_fields", "meta_searchstr_locations",
                                    "meta_searchstr_product_versions"], ColFormatter("", width=15)) \
            .reg_cols_samesettings(
            ["meta_update_date", "meta_deleted", "meta_compliance_rating", "meta_mt_rating", "meta_op_rating",
             "meta_qa_rating", "meta_rating", "meta_profile_completeness"],
            ColFormatter("", width=5, wrap=False)) \
            .reg_cols_samesettings(
            ["compliance_fips_status", "compliance_fips_notes", "compliance_gdpr_status", "compliance_gdpr_notes",
             "compliance_api_status", "compliance_api_notes", "op_guide_status", "op_guide_notes", "op_failover_status",
             "op_failover_notes", "op_horizontal_scalability_status", "op_horizontal_scalability_notes",
             "op_scaling_guide_status", "op_scaling_guide_notes", "op_sla_guide_status", "op_sla_guide_notes",
             "op_metrics_status", "op_metrics_notes", "op_alerts_status", "op_alerts_notes", "op_zero_downtime_status",
             "op_zero_downtime_notes", "op_backup_status", "op_backup_notes", "mt_http_tracing_status",
             "mt_http_tracing_notes", "mt_logging_completeness_status", "mt_logging_completeness_notes",
             "mt_logging_format_status", "mt_logging_format_notes", "mt_logging_storage_status",
             "mt_logging_storage_notes", "mt_logging_sanitization_status", "mt_logging_sanitization_notes",
             "mt_db_anonymisation_status", "mt_db_anonymisation_notes", "qa_manual_tests_status",
             "qa_manual_tests_notes", "qa_unit_tests_status", "qa_unit_tests_notes", "qa_e2e_tests_status",
             "qa_e2e_tests_notes", "qa_perf_tests_status", "qa_perf_tests_notes", "qa_longhaul_tests_status",
             "qa_longhaul_tests_notes", "qa_security_tests_status", "qa_security_tests_notes", "qa_api_tests_status",
             "qa_api_tests_notes", "qa_anonymisation_tests_status", "qa_anonymisation_tests_notes",
             "qa_upgrade_tests_status", "qa_upgrade_tests_notes", "compliance_applicable", "op_applicable",
             "op_safe_restart", "op_safe_delete", "op_safe_redeploy", "mt_applicable", "qa_applicable", "meta_deleted"],
            ColFormatter("", width=15, wrap=False))

        if override_original is True:
            Utility.sheet_settings_map[ComponentVersionModel] = sheet_details
        return sheet_details

    def get_comp_version_obj(self, comp_version: ComponentVersionModel) -> dict:
        try:
            return dict(id=comp_version.id, component=comp_version.component.name, version=comp_version.version.version,
                        category=comp_version.component.category.name,
                        subcategory=comp_version.component.subcategory.name,
                        runtime_type=comp_version.component.runtime_type.name,
                        data_privacy_class=comp_version.component.data_privacy_class.name,
                        vendor=comp_version.component.vendor.name,
                        comments=comp_version.comments,
                        depends_on=super()._get_multi_item_string(comp_version.depends_on.all()),
                        dev_language=super()._get_multi_item_string(comp_version.dev_language),
                        dev_framework=super()._get_multi_item_string(comp_version.dev_framework),
                        dev_database=super()._get_multi_item_string(comp_version.dev_database),
                        dev_orm=super()._get_multi_item_string(comp_version.dev_orm),
                        dev_logging=super()._get_multi_item_string(comp_version.dev_logging),
                        dev_raml=comp_version.dev_raml, dev_repo=comp_version.dev_repo,
                        dev_public_repo=comp_version.dev_public_repo,
                        dev_jira_component=comp_version.dev_jira_component,
                        dev_build_jenkins_job=comp_version.dev_build_jenkins_job, dev_docs=comp_version.dev_docs,
                        dev_public_docs=comp_version.dev_public_docs, dev_commit_link=comp_version.dev_commit_link,
                        dev_api_is_public=comp_version.dev_api_is_public,
                        owner_maintainer=super()._get_printable_person_info(comp_version.owner_maintainer),
                        owner_responsible_qa=super()._get_printable_person_info(comp_version.owner_responsible_qa),
                        owner_product_manager=super()._get_printable_people_info(comp_version.owner_product_manager),
                        owner_program_manager=super()._get_printable_people_info(comp_version.owner_program_manager),
                        owner_escalation_list=super()._get_printable_people_info(comp_version.owner_escalation_list),
                        owner_expert=super()._get_printable_people_info(comp_version.owner_expert),
                        owner_architect=super()._get_printable_people_info(comp_version.owner_architect),
                        meta_product_versions=super()._get_multi_item_string(comp_version.meta_product_versions),
                        compliance_applicable=comp_version.compliance_applicable,
                        compliance_fips_status=comp_version.compliance_fips_status,
                        compliance_fips_notes=comp_version.compliance_fips_notes,
                        compliance_fips_signoff=super()._get_printable_person_info(
                            comp_version.compliance_fips_signoff),
                        compliance_gdpr_status=comp_version.compliance_gdpr_status,
                        compliance_gdpr_notes=comp_version.compliance_gdpr_notes,
                        compliance_gdpr_signoff=super()._get_printable_person_info(
                            comp_version.compliance_gdpr_signoff),
                        compliance_api_status=comp_version.compliance_api_status,
                        compliance_api_notes=comp_version.compliance_api_notes,
                        compliance_api_signoff=super()._get_printable_person_info(comp_version.compliance_api_signoff),
                        op_applicable=comp_version.op_applicable, op_guide_status=comp_version.op_guide_notes,
                        op_guide_signoff=super()._get_printable_person_info(comp_version.op_guide_signoff),
                        op_failover_status=comp_version.op_failover_status,
                        op_failover_notes=comp_version.op_failover_notes,
                        op_failover_signoff=super()._get_printable_person_info(comp_version.op_failover_signoff),
                        op_horizontal_scalability_status=comp_version.op_horizontal_scalability_status,
                        op_horizontal_scalability_notes=comp_version.op_horizontal_scalability_notes,
                        op_horizontal_scalability_signoff=super()._get_printable_person_info(
                            comp_version.op_horizontal_scalability_signoff),
                        op_scaling_guide_status=comp_version.op_scaling_guide_status,
                        op_scaling_guide_notes=comp_version.op_scaling_guide_notes,
                        op_scaling_guide_signoff=super()._get_printable_person_info(
                            comp_version.op_scaling_guide_signoff),
                        op_sla_guide_status=comp_version.op_sla_guide_status,
                        op_sla_guide_notes=comp_version.op_sla_guide_notes,
                        op_sla_guide_signoff=super()._get_printable_person_info(comp_version.op_sla_guide_signoff),
                        op_metrics_status=comp_version.op_metrics_status,
                        op_metrics_notes=comp_version.op_metrics_notes,
                        op_metrics_signoff=super()._get_printable_person_info(comp_version.op_metrics_signoff),
                        op_alerts_status=comp_version.op_alerts_status, op_alerts_notes=comp_version.op_alerts_notes,
                        op_alerts_signoff=super()._get_printable_person_info(comp_version.op_alerts_signoff),
                        op_zero_downtime_status=comp_version.op_zero_downtime_status,
                        op_zero_downtime_notes=comp_version.op_zero_downtime_notes,
                        op_zero_downtime_signoff=super()._get_printable_person_info(
                            comp_version.op_zero_downtime_signoff), op_backup_status=comp_version.op_backup_status,
                        op_backup_notes=comp_version.op_backup_notes,
                        op_backup_signoff=super()._get_printable_person_info(comp_version.op_backup_signoff),
                        op_safe_restart=comp_version.op_safe_restart, op_safe_delete=comp_version.op_safe_delete,
                        op_safe_redeploy=comp_version.op_safe_redeploy, mt_applicable=comp_version.mt_applicable,
                        mt_http_tracing_status=comp_version.mt_http_tracing_status,
                        mt_http_tracing_notes=comp_version.mt_http_tracing_notes,
                        mt_http_tracing_signoff=super()._get_printable_person_info(
                            comp_version.mt_http_tracing_signoff),
                        mt_logging_completeness_status=comp_version.mt_logging_completeness_status,
                        mt_logging_completeness_notes=comp_version.mt_logging_completeness_notes,
                        mt_logging_completeness_signoff=super()._get_printable_person_info(
                            comp_version.mt_logging_completeness_signoff),
                        mt_logging_format_status=comp_version.mt_logging_format_status,
                        mt_logging_format_notes=comp_version.mt_logging_format_notes,
                        mt_logging_format_signoff=super()._get_printable_person_info(
                            comp_version.mt_logging_format_signoff),
                        mt_logging_storage_status=comp_version.mt_logging_storage_status,
                        mt_logging_storage_notes=comp_version.mt_logging_storage_notes,
                        mt_logging_storage_signoff=super()._get_printable_person_info(
                            comp_version.mt_logging_storage_signoff),
                        mt_logging_sanitization_status=comp_version.mt_logging_sanitization_status,
                        mt_logging_sanitization_notes=comp_version.mt_logging_sanitization_notes,
                        mt_logging_sanitization_signoff=super()._get_printable_person_info(
                            comp_version.mt_logging_sanitization_signoff),
                        mt_db_anonymisation_status=comp_version.mt_db_anonymisation_status,
                        mt_db_anonymisation_notes=comp_version.mt_db_anonymisation_notes,
                        mt_db_anonymisation_signoff=super()._get_printable_person_info(
                            comp_version.mt_db_anonymisation_signoff), qa_applicable=comp_version.qa_applicable,
                        qa_manual_tests_status=comp_version.qa_manual_tests_status,
                        qa_manual_tests_signoff=super()._get_printable_person_info(
                            comp_version.qa_manual_tests_signoff),
                        qa_unit_tests_status=comp_version.qa_unit_tests_status,
                        qa_unit_tests_notes=comp_version.qa_unit_tests_notes,
                        qa_unit_tests_signoff=super()._get_printable_person_info(comp_version.qa_unit_tests_signoff),
                        qa_e2e_tests_status=comp_version.qa_e2e_tests_status,
                        qa_e2e_tests_notes=comp_version.qa_e2e_tests_notes,
                        qa_e2e_tests_signoff=super()._get_printable_person_info(comp_version.qa_e2e_tests_signoff),
                        qa_perf_tests_status=comp_version.qa_perf_tests_status,
                        qa_perf_tests_notes=comp_version.qa_perf_tests_notes,
                        qa_perf_tests_signoff=super()._get_printable_person_info(comp_version.qa_perf_tests_signoff),
                        qa_longhaul_tests_status=comp_version.qa_longhaul_tests_status,
                        qa_longhaul_tests_notes=comp_version.qa_longhaul_tests_notes,
                        qa_longhaul_tests_signoff=super()._get_printable_person_info(
                            comp_version.qa_longhaul_tests_signoff),
                        qa_security_tests_status=comp_version.qa_security_tests_status,
                        qa_security_tests_notes=comp_version.qa_security_tests_notes,
                        qa_security_tests_signoff=super()._get_printable_person_info(
                            comp_version.qa_security_tests_signoff),
                        qa_api_tests_status=comp_version.qa_api_tests_status,
                        qa_api_tests_notes=comp_version.qa_api_tests_notes,
                        qa_api_tests_signoff=super()._get_printable_person_info(comp_version.qa_api_tests_signoff),
                        qa_anonymisation_tests_status=comp_version.qa_anonymisation_tests_status,
                        qa_anonymisation_tests_notes=comp_version.qa_anonymisation_tests_notes,
                        qa_anonymisation_tests_signoff=super()._get_printable_person_info(
                            comp_version.qa_anonymisation_tests_signoff),
                        qa_upgrade_tests_status=comp_version.qa_upgrade_tests_status,
                        qa_upgrade_tests_notes=comp_version.qa_upgrade_tests_notes,
                        qa_upgrade_tests_signoff=super()._get_printable_person_info(
                            comp_version.qa_upgrade_tests_signoff))
        except Exception as e:
            raise Exception("Exception {0} occured for record [{1}]".format(e, comp_version))

    def splt_by_latest_version(self, version_objs) -> dict:
        """returns all the objects with their latest version"""
        # returnobj = {
        #     # "latest": [CompVersion1, CompVersion2, ... ]
        #     # "allwithoutlatest": [CompVersion1, CompVersion2, ... ]
        # }

        # ckeys = []
        comps = {}
        for cv in version_objs:
            complst = comps.get(cv.component.name)
            complst = [] if complst is None else complst
            complst.append(cv)
            comps[cv.component.name] = complst

        latest_lst = []
        notlatest_lst = []
        for name, complst in comps.items():
            if len(complst) == 1:
                latest_lst.append(complst[0])
                continue
            latest_time = pytz.utc.localize(datetime.datetime(1790, 1, 1))
            last_idx = 0
            for idx, cv in enumerate(complst):
                if latest_time < cv.version.release_time:
                    last_idx = idx
                    latest_time = cv.version.release_time
            latest_lst.append(complst[last_idx])
            for idx, cv in enumerate(complst):
                if idx != last_idx:
                    notlatest_lst.append(cv)

        latest_comp_ver_objs = []
        notlatest_comp_ver_objs = []
        for comp_version in latest_lst:
            latest_comp_ver_objs.append(self.get_comp_version_obj(comp_version))

        for comp_version in notlatest_lst:
            notlatest_comp_ver_objs.append(self.get_comp_version_obj(comp_version))

        return dict(latest=latest_comp_ver_objs, allwithoutlatest=notlatest_comp_ver_objs)

    def export(self):
        """ Finds and segregates out component versions which are latest and which aren't """
        logging.debug('Exporting for [%s]', ComponentVersionModel)
        tmp_comp_ver_objs = ComponentVersionModel.objects.all()
        logging.debug('Received [%d] records', len(tmp_comp_ver_objs))
        objs_dict = self.splt_by_latest_version(tmp_comp_ver_objs)
        del tmp_comp_ver_objs

        comp_ver_sheet_details = ComponentVersionExporter.get_sheet_details()

        comp_ver_sheet_details.set_tablename(Utility.SHEETNM_COMPONENT_VERSIONS + "_wo_latest")
        self.xlsdoc.update_sheet(comp_ver_sheet_details, objs_dict["allwithoutlatest"])

        comp_ver_sheet_details.set_tablename(Utility.SHEETNM_COMPONENT_VERSIONS + "_latest")
        self.xlsdoc.update_sheet(comp_ver_sheet_details, objs_dict["latest"])
        pass


class DefaultExporter(Exporter):
    def __init__(self, xlsdoc):
        super().__init__(xlsdoc)

    @classmethod
    def get_sheet_details(cls, override_original=True) -> TableFormatter:
        sheet_details = TableFormatter(Utility.SHEETNM_DEFAULTS, sheet_pos=20,
                                       sheet_properties=WorksheetProperties(tabColor="8B0000")) \
                            .reg_cols([ColFormatter("name",
                                                  width=30),
                                       ColFormatter("default",
                                                  width=30)])
        if override_original is True:
            Utility.sheet_settings_map[Utility.SHEETNM_DEFAULTS] = sheet_details
        return sheet_details

    def export(self):
        logging.debug('Exporting for [%s]', Utility.SHEETNM_DEFAULTS)
        entries = [
            {"name": "DeploymentEnvironment", "default" :"K8S"},
            {"name": "DeploymentLocationShortName", "default" : "cloud"},
            {"name": "DeploymentLocationName", "default" : "Cloud datacenter"},
            {"name": "ProductFamily", "default" : "My Cloud Product"},
            {"name": "ProductVersion", "default" : "My Cloud Product 1.0"},
            {"name": "Port", "default" : "TBD"},
            {"name": "ApplicationVendor", "default" : "OpenSource"},
            {"name": "DependencyType", "default" : "sync_rw"},
            {"name": "ComponentDataPrivacy", "default" : "Application"},
            {"name": "ComponentRuntime", "default" : "Component"},
            {"name": "ComponentVersion", "default" : "2.0"},
            {"name": "Comments", "default" : "SEED"}
        ]

        self.xlsdoc.update_sheet(DefaultExporter.get_sheet_details(), entries)

class Command(BaseCommand):
    def handle(self, *args, **options):
        print(options)
        logging.basicConfig(
            format='%(asctime)s,%(msecs)d %(levelname)-8s [%(filename)s:%(funcName)s():%(lineno)d] %(message)s',
            datefmt='%Y-%m-%d:%H:%M:%S',
            level=logging.DEBUG)

        try:
            xlsdoc = XlsDocument(options[Utility.CMDOPTION_XLS_FILE], options[Utility.CMDOPTION_OVERWRITE],
                                 enable_registry=True)
        except Exception as e:
            logging.error("FATAL! [%s]", e)
            print(traceback.format_exc())
            print(Utility.usage())
            sys.exit(-1)

        try:
            GenericExporter(xlsdoc, ComponentCategoryModel).export()
            GenericExporter(xlsdoc, ComponentDataPrivacyClassModel).export()
            GenericExporter(xlsdoc, ComponentRuntimeTypeModel).export()
            GenericExporter(xlsdoc, CountryModel).export()
            GenericExporter(xlsdoc, DatabaseVendorModel).export()
            GenericExporter(xlsdoc, DatacenterModel).export()
            GenericExporter(xlsdoc, DeploymentEnvironmentModel).export()
            GenericExporter(xlsdoc, DeploymentLocationClassModel).export()
            GenericExporter(xlsdoc, LoggerModel).export()
            GenericExporter(xlsdoc, ProgrammingLanguageModel).export()
            GenericExporter(xlsdoc, ProductFamilyModel).export()
            GenericExporter(xlsdoc, SoftwareVendorModel).export()
            GenericExporter(xlsdoc, TCPPortModel).export()
            GenericExporter(xlsdoc, ComponentVersionListModel).export()
            FrameworkExporter(xlsdoc).export()
            ORMExporter(xlsdoc).export()
            ProductVersionExporter(xlsdoc).export()
            ComponentSubcategoryExporter(xlsdoc).export()

            ComponentExporter(xlsdoc).export()
            ComponentVersionExporter(xlsdoc).export()
            ComponentDeploymentExporter(xlsdoc).export()

            DefaultExporter(xlsdoc).export()
        except KeyError as e:
            logging.error("FATAL! [%s]", e)
            print(traceback.format_exc())
            print("Failed exporting data to [{0}]".format(options[Utility.CMDOPTION_XLS_FILE]))
            sys.exit(-1)

        print("Completed exporting data to [{0}]".format(options[Utility.CMDOPTION_XLS_FILE]))

    def add_arguments(self, p):
        g = p.add_argument_group('Generic options')
        g.add_argument('-o', '--overwrite', action='store_true', help='Overwrite existing file')

        g = p.add_argument_group('File generating options')
        g.add_argument('-x', '--xls-file', help='XLS file to generate', required=True)
        Utility.register_parser(p)
